#!/usr/bin/env python3
"""
gmail_amounts_to_excel_multi.py
- Per-account OAuth tokens (switch accounts cleanly)
- Uses Desktop OAuth (system browser) and minimal scopes
- Preserves your money-email parsing & Excel styling logic
"""
import base64
import os
import re
import sys
import argparse
import json
import getpass
import imaplib
import email
from email.header import decode_header, make_header
from datetime import datetime, timezone
from email.utils import parsedate_to_datetime
from pathlib import Path

from google.auth.transport.requests import Request
import pandas as pd
from bs4 import BeautifulSoup
from dateutil import tz

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError

# ---- Excel styling imports (openpyxl) ----
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import DataBarRule
from openpyxl.worksheet.table import Table, TableStyleInfo

# ===================== CONFIG =====================
# Scopes: Gmail read-only + OIDC email to tag per-account tokens
SCOPES = [
    "https://www.googleapis.com/auth/gmail.readonly",
    "openid",
    "https://www.googleapis.com/auth/userinfo.email",
]

# Where to store per-account tokens
TOKENS_DIR = Path(__file__).parent / "tokens"
TOKENS_DIR.mkdir(exist_ok=True)

# Credentials file lives at the repository root
CREDS_FILE = Path(__file__).resolve().parents[2] / "credentials.json"

SEARCH_TERMS = ["montant", "amount", "devis", "facture", "quotation", "invoice", "payment", "paid", "balance", "due"]
EXCEL_PATH = "email_amounts.xlsx"

# Currency & amount patterns:
CURRENCY_SYM = r"(?:€|\$|£|د\.?\s?م\.?|MAD|DHS?|Dh?s?)"
NUM = r"(?:\d{1,3}(?:[.,\s]\d{3})*(?:[.,]\d{2})?|\d+(?:[.,]\d{2})?)"
# Allow optional sign before currency or amount
AMOUNT_REGEX = re.compile(
    rf"(?:[-+]?\s*{CURRENCY_SYM}\s*{NUM}|[-+]?{NUM}\s*{CURRENCY_SYM}|"
    rf"[-+]?{NUM}\s*(?:USD|EUR|GBP|MAD)|"
    rf"(?:USD|EUR|GBP|MAD)\s*[-+]?{NUM})",
    re.IGNORECASE
)

# Normalize currency tokens
CURRENCY_NORMALIZE = {
    "€": "EUR", "eur": "EUR",
    "$": "USD", "usd": "USD",
    "£": "GBP", "gbp": "GBP",
    "mad": "MAD", "dhs": "MAD", "dh": "MAD",
    "د.م.": "MAD", "د.م": "MAD", "د م": "MAD",
}

# Heuristic “promotional” keywords (skip if found in subject/body)
PROMO_KEYWORDS = [
    "unsubscribe", "newsletter", "promotion", "promotional", "deal",
    "special offer", "discount", "limited time", "flash sale", "buy now",
    "exclusive", "coupon", "save %", "sale", "clearance"
]

# Email senders to ignore completely (substring match on the domain/email)
BLOCKED_SENDERS = [
    "reddit",
    "discord",
    "naga.com",
]

# Keywords used to roughly categorize what service an email refers to. The
# keywords are intentionally broad so we can match simple variations of the
# service names (e.g. "advertising", "advertisement", "promo").
SERVICE_KEYWORDS = {
    "photography": ["photo", "picture", "photoshoot", "photography"],
    "filming": ["film", "filming", "video", "recording"],
    "advertising": ["advertis", "marketing", "promo", "campaign"],
}


def is_blocked_sender(sender_email: str) -> bool:
    """Return True if sender_email contains any blocked sender term."""
    s = sender_email.lower()
    return any(term in s for term in BLOCKED_SENDERS)


def detect_service(text: str) -> str:
    """Very simple keyword-based classifier for the service mentioned in text."""
    lower = text.lower()
    for service, keywords in SERVICE_KEYWORDS.items():
        if any(k in lower for k in keywords):
            return service
    return "unknown"

# ============== OAuth helpers (multi-account) ==============
def account_token_path(email: str) -> Path:
    safe = email.replace("@", "_at_").replace(".", "_")
    return TOKENS_DIR / f"token-{safe}.json"

def identify_email_from_creds(creds: Credentials) -> str:
    # Prefer id_token email (requires 'openid userinfo.email')
    try:
        if creds and getattr(creds, "id_token", None):
            id_info = json.loads(creds.id_token)
            if isinstance(id_info, dict) and id_info.get("email"):
                return id_info["email"]
    except Exception:
        pass
    # Fallback: ask Gmail profile
    try:
        service = build("gmail", "v1", credentials=creds)
        profile = service.users().getProfile(userId="me").execute()
        return profile.get("emailAddress")
    except Exception:
        return ""

def load_creds_for_account(account: str | None) -> tuple[Credentials, str]:
    """
    If --account is given and token exists, load it.
    Else run a fresh flow, then store token under that account's email.
    """
    # 1) If account hint provided and token exists, load it
    if account:
        p = account_token_path(account)
        if p.exists():
            creds = Credentials.from_authorized_user_file(str(p), SCOPES)
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            return creds, account

    # 2) Otherwise, try to reuse any token in TOKENS_DIR (first valid found)
    for p in TOKENS_DIR.glob("token-*.json"):
        try:
            c = Credentials.from_authorized_user_file(str(p), SCOPES)
            if c and c.valid:
                email_guess = re.sub(r"^token-|\.json$", "", p.name)
                email_guess = email_guess.replace("_at_", "@").replace("_", ".")
                return c, email_guess
            if c and c.expired and c.refresh_token:
                c.refresh(Request())
                email_guess = identify_email_from_creds(c) or "unknown"
                return c, email_guess
        except Exception:
            continue

    # 3) Fall back to a fresh OAuth flow (system browser)
    flow = InstalledAppFlow.from_client_secrets_file(str(CREDS_FILE), SCOPES)
    creds = flow.run_local_server(port=0, prompt="consent")
    email = identify_email_from_creds(creds) or "unknown"
    # Save under the real account email if we got it
    p = account_token_path(email)
    p.write_text(creds.to_json())
    return creds, email

# ============== IMAP helpers (direct login) ==============
def imap_login(email_address: str, password: str) -> imaplib.IMAP4_SSL:
    """Login to Gmail via IMAP using an email address and password/app password."""
    imap = imaplib.IMAP4_SSL("imap.gmail.com")
    imap.login(email_address, password)
    imap.select("INBOX")
    return imap

def imap_list_ids(imap: imaplib.IMAP4_SSL, days: int, extra_query: str | None = None):
    """Return a list of message UIDs matching the Gmail raw query."""
    q = gmail_query(days, extra_query)
    typ, data = imap.search(None, "X-GM-RAW", q)
    if typ != "OK" or not data:
        return []
    return data[0].split()

def imap_fetch(imap: imaplib.IMAP4_SSL, uid: bytes):
    """Fetch a message by UID and return (email.message.Message, gm_id)."""
    typ, data = imap.fetch(uid, "(RFC822 X-GM-MSGID)")
    if typ != "OK" or not data:
        return None, None
    raw = data[0][1]
    msg = email.message_from_bytes(raw)
    gm_match = re.search(rb"X-GM-MSGID\s+(\d+)", data[0][0])
    gm_id = gm_match.group(1).decode() if gm_match else uid.decode()
    return msg, gm_id

def decode_imap_body(msg):
    """Extract text content from an email.message.Message."""
    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            if ctype == "text/plain":
                try:
                    return part.get_payload(decode=True).decode(errors="ignore")
                except Exception:
                    continue
            if ctype == "text/html":
                try:
                    html = part.get_payload(decode=True).decode(errors="ignore")
                    return BeautifulSoup(html, "html.parser").get_text("\n", strip=True)
                except Exception:
                    continue
    else:
        ctype = msg.get_content_type()
        try:
            payload = msg.get_payload(decode=True).decode(errors="ignore")
        except Exception:
            payload = ""
        if ctype == "text/html":
            return BeautifulSoup(payload, "html.parser").get_text("\n", strip=True)
        return payload
# ============== Gmail search & parsing ==============
def gmail_query(days: int, extra_query: str | None = None) -> str:
    or_terms = " OR ".join(f'"{t}"' for t in SEARCH_TERMS)

    exclusions = [
        "-category:promotions",
        "-category:social",
        "-from:(noreply OR no-reply)",
        "-(subject:newsletter OR unsubscribe OR offer OR promotion OR deal OR sale OR discount)",
    ]
    if BLOCKED_SENDERS:
        exclusions.append("-from:(" + " OR ".join(BLOCKED_SENDERS) + ")")

    q = f"({or_terms}) {' '.join(exclusions)}"
    if days is not None and days > 0:
        q += f" newer_than:{days}d"
    if extra_query:
        q += f" {extra_query}"
    return q

def list_message_ids(service, user_id="me", query=None, max_pages=10, page_size=100):
    all_ids = []
    page_token = None
    for _ in range(max_pages):
        resp = service.users().messages().list(
            userId=user_id, q=query, maxResults=page_size, pageToken=page_token
        ).execute()
        msgs = resp.get("messages", [])
        all_ids.extend(m["id"] for m in msgs)
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return all_ids

def get_message(service, msg_id, user_id="me"):
    return service.users().messages().get(userId=user_id, id=msg_id, format="full").execute()

def decode_body(payload):
    def decode_part(body):
        data = body.get("data")
        if not data:
            return ""
        return base64.urlsafe_b64decode(data).decode("utf-8", errors="ignore")

    if "parts" in payload:
        text, html = "", ""
        stack = list(payload.get("parts", []))
        while stack:
            part = stack.pop()
            mime = part.get("mimeType", "")
            if "parts" in part:
                stack.extend(part["parts"])
            body = part.get("body", {})
            if mime == "text/plain" and body:
                text += "\n" + decode_part(body)
            elif mime == "text/html" and body:
                html += "\n" + decode_part(body)
        if text.strip():
            return text
        if html.strip():
            return BeautifulSoup(html, "html.parser").get_text("\n", strip=True)
        return ""
    else:
        mime = payload.get("mimeType", "")
        body = payload.get("body", {})
        content = decode_part(body)
        if mime == "text/html":
            return BeautifulSoup(content, "html.parser").get_text("\n", strip=True)
        return content

def header(headers, name):
    for h in headers:
        if h.get("name", "").lower() == name.lower():
            return h.get("value", "")
    return ""

def normalize_currency(raw):
    raw = raw.strip().lower()
    for token, code in CURRENCY_NORMALIZE.items():
        if token.lower() in raw:
            return code
    return ""

def value_from_amount(raw):
    # Strip out non-numeric characters except separators and sign
    cleaned = re.sub(r"[^0-9.,\s\-+]", "", raw)
    m = re.findall(r"[-+]?\d[\d.,\s]*", cleaned)
    if not m:
        return None
    num = re.sub(r"\s+", "", m[0])
    if "," in num and "." in num:
        last_comma = num.rfind(",")
        last_dot = num.rfind(".")
        if last_comma > last_dot:
            num = num.replace(".", "").replace(",", ".")
        else:
            num = num.replace(",", "")
    else:
        if "," in num and "." not in num:
            parts = num.split(",")
            if len(parts[-1]) == 3 and all(len(p) == 3 for p in parts[1:]):
                num = "".join(parts)
            else:
                num = num.replace(",", ".")
    try:
        return float(num)
    except ValueError:
        return None

def extract_amounts(text):
    results = []
    for m in AMOUNT_REGEX.finditer(text):
        raw = m.group(0)
        cur = normalize_currency(raw) or ""
        val = value_from_amount(raw)
        if val is not None:
            results.append({"raw": raw, "currency": cur, "value": val})
    return results

def parse_from_header(from_header):
    m = re.search(r'"?([^"<]+)"?\s*<([^>]+)>', from_header)
    if m:
        return m.group(1).strip(), m.group(2).strip()
    if "@" in from_header:
        return "", from_header.strip()
    return from_header.strip(), ""

def gmail_link(message_id):
    return f"https://mail.google.com/mail/u/0/#inbox/{message_id}"

def empty_results_df():
    return pd.DataFrame({
        "date": pd.Series(dtype="datetime64[ns]"),
        "sender_name": pd.Series(dtype="object"),
        "sender_email": pd.Series(dtype="object"),
        "subject": pd.Series(dtype="object"),
        "amount_value": pd.Series(dtype="float"),
        "amount_currency": pd.Series(dtype="object"),
        "amount_raw": pd.Series(dtype="object"),
        "message_id": pd.Series(dtype="object"),
        "snippet": pd.Series(dtype="object"),
        "gmail_link": pd.Series(dtype="object"),
        "account_email": pd.Series(dtype="object"),
        "service": pd.Series(dtype="object"),
    })

def load_existing_excel(path):
    if os.path.exists(path):
        try:
            df = pd.read_excel(path)
            df["date"] = pd.to_datetime(df["date"], errors="coerce")
            if "service" not in df.columns:
                df["service"] = ""
            return df
        except Exception:
            pass
    return empty_results_df()

def to_local_tz_naive(dt: datetime):
    try:
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        local = tz.tzlocal()
        return dt.astimezone(local).replace(tzinfo=None)
    except Exception:
        return dt.replace(tzinfo=None)

def is_promotional(text, subject):
    content = f"{subject or ''} {text or ''}".lower()
    return any(k in content for k in PROMO_KEYWORDS)

# ---------- Excel styling helpers ----------
HEADER_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)

# Light fills for different currencies so the Excel sheet is easier to scan
CURRENCY_FILL = {
    "USD": "D9EAD3",   # light green
    "EUR": "D0E0E3",   # light blue
    "GBP": "FCE5CD",   # light orange
    "MAD": "F4CCCC",   # light red/pink
}

def best_fit_width(values, min_w=10, max_w=60):
    try:
        lengths = [len(str(v)) for v in values if v is not None]
        if not lengths:
            return min_w
        return max(min(max(lengths) + 2, min_w), max_w)
    except Exception:
        return min_w

def style_excel(path, sheet_name="Sheet1", make_table=True):
    wb = load_workbook(path)
    ws = wb.active if sheet_name not in wb.sheetnames else wb[sheet_name]
    ws.title = "email_amounts"

    # Freeze header row
    ws.freeze_panes = "A2"

    # Header style
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    # Autofilter range
    last_col_letter = ws.cell(row=1, column=ws.max_column).column_letter
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

    # Column formats & widths
    headers = [c.value for c in ws[1]]
    col_map = {h: i+1 for i, h in enumerate(headers)}

    # Date format
    if "date" in col_map:
        col = col_map["date"]
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = "yyyy-mm-dd hh:mm"

    # Amount number format (2 decimals with thousands)
    if "amount_value" in col_map:
        col = col_map["amount_value"]
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = '#,##0.00'

        # Data bar on amount_value (use min/max types exactly as required by openpyxl)
        data_range = f"{ws.cell(row=2, column=col).coordinate}:{ws.cell(row=ws.max_row, column=col).coordinate}"
        if ws.max_row >= 2:
            rule = DataBarRule(start_type="min", start_value=None,
                               end_type="max", end_value=None,
                               color="63BE7B", showValue=True)
            ws.conditional_formatting.add(data_range, rule)

    # Wrap text for subject/snippet
    for name in ["subject", "snippet"]:
        if name in col_map:
            col = col_map[name]
            for row in range(2, ws.max_row + 1):
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    # Give currency cells a light fill so users can visually separate currencies
    if "amount_currency" in col_map:
        col = col_map["amount_currency"]
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=col)
            color = CURRENCY_FILL.get(str(cell.value).upper())
            if color:
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Borders + column widths
    data = list(ws.iter_rows(values_only=True))
    for col_idx in range(1, ws.max_column + 1):
        col_values = [row[col_idx - 1] for row in data] if data else []
        width = best_fit_width(col_values)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).border = THIN_BORDER

    # Optional: Excel Table with nice banded style
    if make_table and ws.max_row >= 1 and ws.max_column >= 1:
        ref = f"A1:{last_col_letter}{ws.max_row}"
        if ws._tables:
            ws._tables.clear()
        tbl = Table(displayName="EmailAmounts", ref=ref)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tbl.tableStyleInfo = style
        ws.add_table(tbl)

    wb.save(path)


def write_totals_sheet(path: str | Path, df: pd.DataFrame) -> None:
    """Write a totals sheet (sum of amounts per currency) to the workbook."""
    totals = (
        df.groupby("amount_currency")["amount_value"].sum().reset_index()
        if not df.empty
        else pd.DataFrame(columns=["amount_currency", "amount_value"])
    )

    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        totals.to_excel(writer, sheet_name="totals", index=False)

    wb = load_workbook(path)
    ws = wb["totals"]

    # Style header
    if ws.max_row >= 1:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

    # Format amount column and borders
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        for col in range(1, ws.max_column + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER

    last_col_letter = ws.cell(row=1, column=ws.max_column).column_letter
    ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

    wb.save(path)


# ---------- Simple analytics ----------
def most_profitable_service(df: pd.DataFrame) -> str | None:
    """Return the service with the highest total amount in the dataframe."""
    if df.empty or "service" not in df.columns:
        return None
    totals = df.groupby("service")["amount_value"].sum().sort_values(ascending=False)
    if totals.empty:
        return None
    return totals.index[0]


# ===================== SCRAPER ENTRY POINT =====================
def run_scraper(days=180, query=None, take=None, pick="max", account=None, email=None, password=None):
    """Run the scraper logic and return a DataFrame with the results."""
    try:
        if email:
            if not password:
                raise ValueError("Password is required when using direct IMAP login")
            imap = imap_login(email, password)
            msg_ids = imap_list_ids(imap, days, query)
            if take:
                msg_ids = msg_ids[:take]

            df = load_existing_excel(EXCEL_PATH)
            existing_ids = set(df["message_id"].astype(str).tolist())

            rows = []
            for uid in msg_ids:
                msg, gm_id = imap_fetch(imap, uid)
                if not msg or gm_id in existing_ids:
                    continue
                subject = str(make_header(decode_header(msg.get("Subject", ""))))
                from_h = str(make_header(decode_header(msg.get("From", ""))))
                date_h = msg.get("Date", "")
                body_text = decode_imap_body(msg)
                if is_promotional(body_text, subject):
                    continue
                snippet = body_text[:200]
                sender_name, sender_email = parse_from_header(from_h)
                if is_blocked_sender(sender_email):
                    continue

                try:
                    dt = parsedate_to_datetime(date_h)
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                except Exception:
                    dt = datetime.utcnow().replace(tzinfo=timezone.utc)
                dt_local = to_local_tz_naive(dt)

                text_for_amounts = f"{subject}\n{body_text}"
                service_type = detect_service(text_for_amounts)
                amts = extract_amounts(text_for_amounts)
                if not amts:
                    continue

                if pick == "first":
                    chosen = [amts[0]]
                elif pick == "max":
                    chosen = [max(amts, key=lambda x: x["value"])]
                else:
                    chosen = amts

                for a in chosen:
                    rows.append({
                        "date": dt_local.replace(microsecond=0),
                        "sender_name": sender_name,
                        "sender_email": sender_email,
                        "subject": subject,
                        "amount_value": a["value"],
                        "amount_currency": a["currency"],
                        "amount_raw": a["raw"],
                        "message_id": gm_id,
                        "snippet": snippet,
                        "gmail_link": gmail_link(gm_id),
                        "account_email": email,
                        "service": service_type,
                    })

            if rows:
                out = pd.DataFrame(rows)
                out["date"] = pd.to_datetime(out["date"], errors="coerce")
                merged = pd.concat([df, out], ignore_index=True)
                merged.drop_duplicates(subset=["message_id", "amount_raw"], keep="first", inplace=True)
                merged["date"] = pd.to_datetime(merged["date"]).dt.tz_localize(None)
                merged.sort_values(by="date", ascending=False, inplace=True)
                merged.to_excel(EXCEL_PATH, index=False)
                style_excel(EXCEL_PATH)
                write_totals_sheet(EXCEL_PATH, merged)
            else:
                merged = df
            imap.logout()
        else:
            creds, account_email = load_creds_for_account(account)
            service = build("gmail", "v1", credentials=creds)

            q = gmail_query(days, query)
            msg_ids = list_message_ids(service, query=q)
            if take:
                msg_ids = msg_ids[:take]

            df = load_existing_excel(EXCEL_PATH)
            existing_ids = set(df["message_id"].astype(str).tolist())

            rows = []
            for mid in msg_ids:
                if str(mid) in existing_ids:
                    continue
                msg = get_message(service, mid)
                payload = msg.get("payload", {})
                headers = payload.get("headers", [])
                subject = header(headers, "Subject")
                from_h = header(headers, "From")
                date_h = header(headers, "Date")
                snippet = msg.get("snippet", "")

                sender_name, sender_email = parse_from_header(from_h)
                if is_blocked_sender(sender_email):
                    continue

                try:
                    dt = parsedate_to_datetime(date_h)
                    if dt.tzinfo is None:
                        dt = dt.replace(tzinfo=timezone.utc)
                except Exception:
                    dt = datetime.utcnow().replace(tzinfo=timezone.utc)
                dt_local = to_local_tz_naive(dt)

                body_text = decode_body(payload)
                if is_promotional(body_text, subject):
                    continue

                text_for_amounts = f"{subject}\n{body_text}"
                service_type = detect_service(text_for_amounts)
                amts = extract_amounts(text_for_amounts)
                if not amts:
                    continue

                if pick == "first":
                    chosen = [amts[0]]
                elif pick == "max":
                    chosen = [max(amts, key=lambda x: x["value"])]
                else:
                    chosen = amts  # all

                for a in chosen:
                    rows.append({
                        "date": dt_local.replace(microsecond=0),
                        "sender_name": sender_name,
                        "sender_email": sender_email,
                        "subject": subject,
                        "amount_value": a["value"],
                        "amount_currency": a["currency"],
                        "amount_raw": a["raw"],
                        "message_id": mid,
                        "snippet": snippet,
                        "gmail_link": gmail_link(mid),
                        "account_email": account_email,
                        "service": service_type,
                    })

            if rows:
                out = pd.DataFrame(rows)
                out["date"] = pd.to_datetime(out["date"], errors="coerce")
                merged = pd.concat([df, out], ignore_index=True)
                merged.drop_duplicates(subset=["message_id", "amount_raw"], keep="first", inplace=True)
                merged["date"] = pd.to_datetime(merged["date"]).dt.tz_localize(None)
                merged.sort_values(by="date", ascending=False, inplace=True)
                merged.to_excel(EXCEL_PATH, index=False)
                style_excel(EXCEL_PATH)
                write_totals_sheet(EXCEL_PATH, merged)
            else:
                merged = df
        top_service = most_profitable_service(merged)
        if top_service:
            print(f"Most profitable service: {top_service}")
        return merged
    except HttpError as e:
        raise e
    except Exception as e:
        raise e


def main():
    parser = argparse.ArgumentParser(description="Extract email amounts into Excel (multi-account safe)")
    parser.add_argument("--days", type=int, default=180, help="Search in the last N days. 0 = all mail.")
    parser.add_argument("--query", type=str, default=None, help="Extra Gmail search query")
    parser.add_argument("--take", type=int, default=None, help="Stop after processing this many messages (testing).")
    parser.add_argument("--pick", choices=["first", "max", "all"], default="max",
                        help="If multiple amounts in one email: first, max, or all (default: max).")
    parser.add_argument("--account", type=str, default=None,
                        help="Email address of the Google account to use (if you have multiple tokens).")
    parser.add_argument("--email", type=str, default=None,
                        help="Use direct IMAP login with this email instead of OAuth.")
    parser.add_argument("--password", type=str, default=None,
                        help="Password or app password for IMAP login (will prompt if omitted).")
    args = parser.parse_args()

    if not args.email and not CREDS_FILE.exists():
        print(f"Google API credentials file not found: {CREDS_FILE}")
        sys.exit(1)

    if not args.email:
        choice = input(
            "Choose login method: [1] Automatic via browser link (default) or [2] Manual email/password: "
        ).strip()
        if choice == "2":
            args.email = input("Gmail address: ").strip()
            args.password = getpass.getpass("Gmail password or app password: ")

    try:
        run_scraper(**vars(args))
    except HttpError as e:
        print(f"Gmail API error: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
